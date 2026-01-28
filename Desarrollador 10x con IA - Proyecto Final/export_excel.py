import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from models import WorkItem

def exportar_excel(work_items: list[WorkItem], archivo: str, sprint_nombre: str):
    filas = []

    for wi in work_items:
        filas.append([
            wi.work_item_type,
            wi.id,
            wi.title,
            wi.description,
            wi.state,
            wi.iteration_path,
            wi.parent
        ])

    df = pd.DataFrame(
        filas,
        columns=[
            "Work Item Type",
            "ID",
            "Title",
            "Description",
            "State",
            "Iteration Path",
            "Parent"
        ]
    )

    df.to_excel(archivo, index=False, startrow=2)

    wb = load_workbook(archivo)
    ws = wb.active

    ws["A1"] = sprint_nombre
    ws["A1"].font = Font(bold=True, size=14)

    wb.save(archivo)